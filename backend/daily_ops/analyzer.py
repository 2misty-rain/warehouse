"""
数据分析模块
功能：
1. 处理当天原始数据，筛选异常用户
2. 生成异常明细和异常汇总
3. 对比昨天和今天的设备状态，生成状态变化表

设计说明：
- 所有函数接收数据源参数，支持未来切换为数据库查询
- 不使用全局状态，便于 Web 服务并发调用
"""

import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta
from typing import Optional, List, Tuple, Dict, Any

from . import api_client

logger = logging.getLogger(__name__)

# 颜色定义
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# 状态分类
STATUS_HAS_USER = ['正常', '长期卧床']
STATUS_NO_USER = ['长期无人', '睡眠时间断电']

# 需要忽略的机构列表
IGNORE_INSTITUTIONS = ['溪翁镇社会福利中心', '溪翁庄镇社会福利中心']

# 需要屏蔽的设备号列表
IGNORE_DEVICES = ['CAA241100120', 'CAA231200292', 'DAA251200001', 'CAA241100125']


def get_institution(device_id: str, device_mapping_cache: dict) -> str:
    """根据设备号从映射缓存查询归属机构简称"""
    if device_mapping_cache and device_id in device_mapping_cache:
        cache_entry = device_mapping_cache[device_id]
        if isinstance(cache_entry, dict):
            return cache_entry.get("organizationName", "")
    return ""


def should_ignore_device(device_id: str, device_mapping_cache: dict) -> bool:
    """判断是否应该忽略该设备"""
    institution = get_institution(device_id, device_mapping_cache)
    return institution in IGNORE_INSTITUTIONS


def prepare_dataframe(file_path: str) -> pd.DataFrame:
    """读取 Excel 文件，取消合并单元格，处理表头"""
    wb = load_workbook(file_path)
    ws = wb.active

    merged_cells = list(ws.merged_cells.ranges)
    for merged in merged_cells:
        ws.unmerge_cells(str(merged))

    temp_path = file_path.replace('.xlsx', '_temp.xlsx')
    wb.save(temp_path)
    wb.close()

    df = pd.read_excel(temp_path, header=None)

    header_row1 = df.iloc[0].tolist()
    header_row2 = df.iloc[1].tolist()

    new_columns = []
    for i in range(len(header_row1)):
        h1 = header_row1[i] if pd.notna(header_row1[i]) else None
        h2 = header_row2[i] if pd.notna(header_row2[i]) else None
        if h1 and h2:
            new_columns.append(f"{h1}_{h2}")
        elif h1:
            new_columns.append(h1)
        elif h2:
            new_columns.append(h2)
        else:
            new_columns.append(f"列{i + 1}")

    df.columns = new_columns
    df = df.iloc[2:].reset_index(drop=True)
    df = df.dropna(subset=['设备号'])
    os.remove(temp_path)
    return df


# 需要动态匹配的列定义
REQUIRED_COLUMNS = {
    '睡眠时长(小时)': ['睡眠时长', '睡眠时间'],
    '夜间离床次数': ['离床次数', '夜间离床', '离床'],
    '呼吸过速次数': ['呼吸过速', '呼吸速'],
    '呼吸过缓次数': ['呼吸过缓', '呼吸缓'],
    '呼吸暂停次数': ['呼吸暂停', '呼吸停'],
    '心动过速次数': ['心动过速', '心率过速'],
    '心动过缓次数': ['心动过缓', '心率过缓']
}


def resolve_column_mapping(df: pd.DataFrame) -> dict:
    """动态匹配列名，返回 {标准名: 实际列名} 的映射"""
    col_map = {}
    for target_col, keywords in REQUIRED_COLUMNS.items():
        for df_col in df.columns:
            for keyword in keywords:
                if keyword in str(df_col):
                    col_map[target_col] = df_col
                    break
            if target_col in col_map:
                break
    return col_map


def apply_column_mapping(df: pd.DataFrame, col_map: dict) -> pd.DataFrame:
    """根据列映射将实际列转换为数值型的标准列"""
    df = df.copy()
    for target_col, src_col in col_map.items():
        if src_col in df.columns:
            df[target_col] = pd.to_numeric(df[src_col], errors='coerce').fillna(0)
        else:
            df[target_col] = 0
    return df


def filter_abnormal_users(df: pd.DataFrame, today_str: str, device_mapping_cache: dict) -> pd.DataFrame:
    """筛选当天异常用户（包含有作息的长期卧床设备）"""
    df['日期_str'] = df['日期'].astype(str).str[:10]
    df_today = df[df['日期_str'] == today_str].copy()

    if df_today.empty:
        return pd.DataFrame()

    # 过滤忽略的机构和设备
    mask_keep = df_today['设备号'].apply(
        lambda did: not should_ignore_device(str(did), device_mapping_cache) and str(did) not in IGNORE_DEVICES
    )
    df_today = df_today[mask_keep].copy()

    # 纳入"正常" + "长期卧床"中有作息的设备
    df_valid = df_today[df_today['用户设备状态'].isin(['正常', '长期卧床'])].copy()
    if df_valid.empty:
        return pd.DataFrame()

    # 动态列映射并转换（需要先映射才能判断长期卧床设备是否有作息）
    col_map = resolve_column_mapping(df_valid)
    df_valid = apply_column_mapping(df_valid, col_map)

    # 长期卧床设备只保留有作息的（睡眠时长 > 0）
    long_term_mask = df_valid['用户设备状态'] == '长期卧床'
    if long_term_mask.any():
        has_activity = df_valid['睡眠时长(小时)'] > 0
        df_valid = df_valid[~long_term_mask | has_activity].copy()

    if df_valid.empty:
        return pd.DataFrame()

    # 标记异常
    df_valid['睡眠时长异常'] = df_valid['睡眠时长(小时)'] < 4
    df_valid['夜间离床异常'] = df_valid['夜间离床次数'] >= 4
    df_valid['呼吸过速异常'] = df_valid['呼吸过速次数'] > 0
    df_valid['呼吸过缓异常'] = df_valid['呼吸过缓次数'] > 0
    df_valid['呼吸暂停异常'] = df_valid['呼吸暂停次数'] > 0
    df_valid['心动过速异常'] = df_valid['心动过速次数'] > 0
    df_valid['心动过缓异常'] = df_valid['心动过缓次数'] > 0

    abnormal_mask = (
        df_valid['睡眠时长异常'] |
        df_valid['夜间离床异常'] |
        df_valid['呼吸过速异常'] |
        df_valid['呼吸过缓异常'] |
        df_valid['呼吸暂停异常'] |
        df_valid['心动过速异常'] |
        df_valid['心动过缓异常']
    )
    return df_valid[abnormal_mask].copy()


def generate_abnormal_summary(df_abnormal: pd.DataFrame, today_str: str,
                               device_mapping_cache: dict, token: str) -> List[dict]:
    """生成异常汇总数据"""
    summary_rows = []
    yesterday_str = (datetime.strptime(today_str, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    event_time = f"{today_str[5:].replace('-', '/')}睡眠期间"

    for _, row in df_abnormal.iterrows():
        device_id = str(row.get('设备号', ''))
        name = row.get('姓名', '')
        institution = get_institution(device_id, device_mapping_cache)

        sleep_duration = row.get('睡眠时长(小时)', 0)
        bed_exit_count = row.get('夜间离床次数', 0)
        tachypnea = row.get('呼吸过速次数', 0)
        bradypnea = row.get('呼吸过缓次数', 0)
        apnea = row.get('呼吸暂停次数', 0)
        tachycardia = row.get('心动过速次数', 0)
        bradycardia = row.get('心动过缓次数', 0)

        sleep_start = row.get('睡眠详情_就寝时间', '') or row.get('入睡时间', '')
        sleep_end = row.get('起床时间', '')

        # 睡眠问题
        if pd.notna(sleep_duration) and sleep_duration == 0:
            summary_rows.append({
                '日期': today_str, '机构': institution, '设备号': device_id,
                '姓名': name, '问题分类': '睡眠状态异常',
                '事件发生时间': event_time, '事件记录': '护理端不显示睡眠记录'
            })
        elif pd.notna(sleep_duration) and 0 < sleep_duration < 4:
            sleep_record = f"{sleep_start}-{sleep_end},睡眠{sleep_duration}小时"
            summary_rows.append({
                '日期': today_str, '机构': institution, '设备号': device_id,
                '姓名': name, '问题分类': '睡眠过少',
                '事件发生时间': event_time, '事件记录': sleep_record
            })

        # 离床问题
        if pd.notna(bed_exit_count) and bed_exit_count >= 4:
            events = api_client.query_bed_exit_events(device_id, today_str, today_str, token, device_mapping_cache)
            events_filtered = api_client.filter_bed_exit_events_for_today(events, today_str)
            bed_exit_detail = api_client.format_bed_exit_events(events_filtered)
            summary_rows.append({
                '日期': today_str, '机构': institution, '设备号': device_id,
                '姓名': name, '问题分类': '多次离床',
                '事件发生时间': event_time, '事件记录': bed_exit_detail
            })

        # 体征异常
        vital_signs = []
        if pd.notna(tachypnea) and tachypnea > 0:
            vital_signs.append(f"呼吸过速{int(tachypnea)}次")
        if pd.notna(bradypnea) and bradypnea > 0:
            vital_signs.append(f"呼吸过缓{int(bradypnea)}次")
        if pd.notna(apnea) and apnea > 0:
            vital_signs.append(f"呼吸暂停{int(apnea)}次")
        if pd.notna(tachycardia) and tachycardia > 0:
            vital_signs.append(f"心动过速{int(tachycardia)}次")
        if pd.notna(bradycardia) and bradycardia > 0:
            vital_signs.append(f"心动过缓{int(bradycardia)}次")

        if vital_signs:
            summary_rows.append({
                '日期': today_str, '机构': institution, '设备号': device_id,
                '姓名': name, '问题分类': '体征异常',
                '事件发生时间': event_time, '事件记录': '、'.join(vital_signs)
            })

    return summary_rows


def detect_status_changes(df_yesterday: pd.DataFrame, df_today: pd.DataFrame,
                          device_mapping_cache: dict) -> List[Tuple]:
    """检测昨天→今天的设备状态变化（有人↔无人切换）"""
    cols_needed = ['设备号', '用户设备状态']
    df_y = df_yesterday[cols_needed].copy()
    df_y['日期'] = '昨天'
    df_t = df_today[cols_needed].copy()
    df_t['日期'] = '今天'

    df_combined = pd.concat([df_y, df_t], ignore_index=True)
    df_sorted = df_combined.sort_values(by=['设备号', '日期'], ascending=[True, True])

    result_rows = []
    for device_id, group in df_sorted.groupby('设备号'):
        if len(group) < 2:
            continue
        yesterday_status = group[group['日期'] == '昨天']['用户设备状态'].values
        today_status = group[group['日期'] == '今天']['用户设备状态'].values
        if len(yesterday_status) == 0 or len(today_status) == 0:
            continue

        prev_status = yesterday_status[0]
        current_status = today_status[0]
        prev_has_user = prev_status in STATUS_HAS_USER
        current_has_user = current_status in STATUS_HAS_USER

        if prev_has_user != current_has_user:
            if should_ignore_device(device_id, device_mapping_cache):
                continue
            today_row = df_today[df_today['设备号'] == device_id].iloc[0].copy()
            institution = get_institution(device_id, device_mapping_cache)
            result_rows.append((today_row, prev_status, current_status, institution))

    return result_rows


def save_to_excel(df_abnormal: pd.DataFrame, summary_rows: List[dict],
                  status_change_rows: List[Tuple], output_file: str,
                  device_mapping_cache: dict, token: str, room_info_cache: dict = None):
    """保存结果到 Excel，包含 3 个 Sheet"""
    from openpyxl import Workbook

    if room_info_cache is None:
        room_info_cache = {}

    wb = Workbook()

    # ========== Sheet1: 异常明细 ==========
    ws1 = wb.active
    ws1.title = "异常明细"

    if not df_abnormal.empty:
        exclude_cols = ['日期_str', '睡眠时长异常', '夜间离床异常', '呼吸过速异常',
                        '呼吸过缓异常', '呼吸暂停异常', '心动过速异常', '心动过缓异常']
        display_cols = [c for c in df_abnormal.columns if c not in exclude_cols]

        if '设备号' in display_cols:
            dev_idx = display_cols.index('设备号')
            display_cols.insert(dev_idx + 1, '机构')
            display_cols.insert(dev_idx + 2, '房间号')

        for c_idx, col_name in enumerate(display_cols, 1):
            ws1.cell(row=1, column=c_idx, value=col_name)

        abnormal_col_names = {
            '睡眠时长(小时)', '夜间离床次数', '呼吸过速次数', '呼吸过缓次数',
            '呼吸暂停次数', '心动过速次数', '心动过缓次数'
        }

        for r_idx, (_, row) in enumerate(df_abnormal.iterrows(), 1):
            device_id = str(row.get('设备号', ''))
            institution = get_institution(device_id, device_mapping_cache)
            room_info = api_client.get_room_info(device_id, token, room_info_cache)

            for c_idx, col_name in enumerate(display_cols, 1):
                if col_name == '机构':
                    cell = ws1.cell(row=r_idx + 1, column=c_idx, value=institution)
                elif col_name == '房间号':
                    cell = ws1.cell(row=r_idx + 1, column=c_idx, value=room_info)
                else:
                    value = row.get(col_name)
                    cell = ws1.cell(row=r_idx + 1, column=c_idx, value=value)

                if col_name in abnormal_col_names:
                    marker_col = col_name.replace('次数', '异常').replace('(小时)', '异常')
                    if marker_col in row.index and row[marker_col]:
                        cell.fill = YELLOW_FILL

    # ========== Sheet2: 异常汇总 ==========
    ws2 = wb.create_sheet(title="异常汇总")

    if summary_rows:
        headers = ['日期', '机构', '房间号', '设备号', '姓名', '问题分类', '事件发生时间', '事件记录']
        for c_idx, header in enumerate(headers, 1):
            ws2.cell(row=1, column=c_idx, value=header)

        for r_idx, row_data in enumerate(summary_rows, 1):
            for c_idx, header in enumerate(headers, 1):
                if header == '机构':
                    value = row_data.get('机构', row_data.get('机构名称', ''))
                elif header == '房间号':
                    did = row_data.get('设备号', '')
                    value = api_client.get_room_info(str(did), token, room_info_cache) if did else ''
                else:
                    value = row_data.get(header, '')
                cell = ws2.cell(row=r_idx + 1, column=c_idx, value=value)

        for row in ws2.iter_rows(min_row=2, max_row=len(summary_rows) + 1, min_col=8, max_col=8):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws2.row_dimensions[cell.row].height = 15 * (str(cell.value).count('\n') + 1) if cell.value else 15

    # ========== Sheet3: 状态变化设备 ==========
    ws3 = wb.create_sheet(title="状态变化设备")

    if status_change_rows:
        data_rows = []
        for today_row, prev_status, current_status, institution in status_change_rows:
            row_dict = today_row.to_dict()
            row_dict['前一天用户设备状态'] = prev_status
            row_dict['今天用户设备状态'] = current_status
            row_dict['机构'] = institution
            device_id = row_dict.get('设备号', '')
            room_info = api_client.get_room_info(str(device_id), token, room_info_cache) if device_id else ''
            row_dict['房间号'] = room_info
            data_rows.append(row_dict)

        result_df = pd.DataFrame(data_rows)

        col_order = []
        for col in result_df.columns:
            if col in ['机构', '房间号']:
                continue
            col_order.append(col)
            if col == '设备号':
                col_order.append('机构')
                col_order.append('房间号')

        result_df = result_df[[c for c in col_order if c in result_df.columns]]

        for c_idx, col_name in enumerate(result_df.columns, 1):
            ws3.cell(row=1, column=c_idx, value=col_name)

        status_cols = ['前一天用户设备状态', '今天用户设备状态']
        for r_idx, row in enumerate(result_df.values, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws3.cell(row=r_idx + 1, column=c_idx, value=value)
                col_name = result_df.columns[c_idx - 1]
                if col_name in status_cols:
                    cell.fill = ORANGE_FILL

    wb.save(output_file)
    wb.close()


def analyze_today(today_str: str, data_dir: str, token: str = None,
                  device_mapping_cache: dict = None, room_info_cache: dict = None) -> dict:
    """
    分析指定日期的数据

    Args:
        today_str: 日期字符串 YYYY-MM-DD
        data_dir: 数据文件存放目录（日期/ 目录的父目录）
        token: API 认证 token
        device_mapping_cache: 设备映射缓存
        room_info_cache: 房间信息缓存

    Returns:
        dict: 分析结果，包含状态和统计信息
    """
    if device_mapping_cache is None:
        device_mapping_cache = {}
    if room_info_cache is None:
        room_info_cache = {}

    today_folder = os.path.join(data_dir, today_str)
    today_file = os.path.join(today_folder, f"原始数据_{today_str}.xlsx")

    yesterday = (datetime.strptime(today_str, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    yesterday_folder = os.path.join(data_dir, yesterday)
    yesterday_file = os.path.join(yesterday_folder, f"原始数据_{yesterday}.xlsx")

    if not os.path.exists(today_file):
        return {"success": False, "error": f"找不到今天的原始数据文件：{today_file}"}

    # 获取 token
    if token is None:
        token = api_client.login()

    logger.info(f"分析日期: {today_str}, 缓存设备数: {len(device_mapping_cache)}")

    # 准备数据
    df_today = prepare_dataframe(today_file)
    logger.info(f"今天数据: {len(df_today)} 条")

    # 筛选异常用户
    df_abnormal = filter_abnormal_users(df_today, today_str, device_mapping_cache)
    logger.info(f"发现 {len(df_abnormal)} 个异常用户")

    # 生成异常汇总
    summary_rows = generate_abnormal_summary(df_abnormal, today_str, device_mapping_cache, token)
    logger.info(f"生成 {len(summary_rows)} 条汇总记录")

    # 检测状态变化
    status_change_rows = []
    if os.path.exists(yesterday_file):
        df_yesterday = prepare_dataframe(yesterday_file)
        logger.info(f"昨天数据: {len(df_yesterday)} 条")
        status_change_rows = detect_status_changes(df_yesterday, df_today, device_mapping_cache)
        logger.info(f"发现 {len(status_change_rows)} 个状态变化的设备")

    # 保存 Excel
    output_file = os.path.join(today_folder, f"异常汇总_{today_str}.xlsx")
    save_to_excel(df_abnormal, summary_rows, status_change_rows, output_file,
                  device_mapping_cache, token, room_info_cache)

    # 统计有效设备数
    valid_device_count = 0
    if os.path.exists(today_file):
        df_original = pd.read_excel(today_file)
        if '用户设备状态' in df_original.columns:
            valid_device_count = len(df_original[df_original['用户设备状态'] == '正常'])

    # 统计各类异常数量
    sleep_too_little = sum(1 for r in summary_rows if r.get('问题分类') == '睡眠过少')
    multiple_bed_exit = sum(1 for r in summary_rows if r.get('问题分类') == '多次离床')
    sleep_abnormal = sum(1 for r in summary_rows if r.get('问题分类') == '睡眠状态异常')
    vital_abnormal = sum(1 for r in summary_rows if r.get('问题分类') == '体征异常')

    # 生成文本摘要
    summary_text = f"【{today_str}】\n"
    summary_text += f"有效设备{valid_device_count}台\n"
    summary_text += f"异常明细{len(df_abnormal)}条\n"
    summary_text += f"异常汇总{len(summary_rows)}条，其中：\n"
    summary_text += f"睡眠过少{sleep_too_little}例，多次离床{multiple_bed_exit}例，"
    summary_text += f"睡眠状态异常{sleep_abnormal}例，体征异常{vital_abnormal}例\n"
    summary_text += f"状态变化设备{len(status_change_rows)}台"

    # 保存文本报告
    txt_file = os.path.join(today_folder, f"汇总报告_{today_str}.txt")
    with open(txt_file, 'w', encoding='utf-8') as f:
        f.write(summary_text)

    return {
        "success": True,
        "date": today_str,
        "output_file": output_file,
        "txt_file": txt_file,
        "summary_text": summary_text,
        "stats": {
            "valid_devices": valid_device_count,
            "abnormal_devices": len(df_abnormal),
            "total_summary": len(summary_rows),
            "sleep_too_little": sleep_too_little,
            "multiple_bed_exit": multiple_bed_exit,
            "sleep_abnormal": sleep_abnormal,
            "vital_abnormal": vital_abnormal,
            "status_changes": len(status_change_rows),
        },
        "summary_rows": summary_rows,
        "status_change_count": len(status_change_rows),
    }


def get_device_daily_status(today_str: str, data_dir: str, device_mapping_cache: dict = None) -> List[dict]:
    """
    预留：获取设备每日状态列表（用于设备每日状态运维窗口）

    当前从 Excel 文件中读取，未来可从数据库直连查询。
    """
    if device_mapping_cache is None:
        device_mapping_cache = {}

    today_folder = os.path.join(data_dir, today_str)
    today_file = os.path.join(today_folder, f"原始数据_{today_str}.xlsx")

    if not os.path.exists(today_file):
        return []

    df = prepare_dataframe(today_file)

    # 应用动态列映射，使标准列名可用
    col_map = resolve_column_mapping(df)
    df = apply_column_mapping(df, col_map)

    devices = []
    for _, row in df.iterrows():
        device_id = str(row.get('设备号', ''))
        if not device_id:
            continue
        if str(device_id) in IGNORE_DEVICES:
            continue
        if should_ignore_device(device_id, device_mapping_cache):
            continue

        sleep_val = row.get('睡眠时长(小时)', 0)
        bed_val = row.get('夜间离床次数', 0)

        devices.append({
            'device_id': device_id,
            'name': str(row.get('姓名', '')),
            'status': str(row.get('用户设备状态', '')),
            'institution': get_institution(device_id, device_mapping_cache),
            'sleep_duration': float(sleep_val) if pd.notna(sleep_val) else 0.0,
            'bed_exit_count': int(bed_val) if pd.notna(bed_val) else 0,
        })

    return devices
